import json
import threading
import time
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

import invoice_agent.web as web
import invoice_agent.pdf_export as pdf_export
from invoice_agent.config import load_agent_config
from invoice_agent.cli import _trip_info_from_args, build_parser
from invoice_agent.company_reimbursement import (
    amount_to_chinese_upper,
    build_company_reimbursement_data,
    company_template_path,
    write_company_workbook,
)
from invoice_agent.extractor import extract_fields_from_text
from invoice_agent.excel import build_preview, write_workbook
from invoice_agent.models import ExpenseRecord, ParsedDocument, TripInfo
from invoice_agent.ocr import PaddleAsyncOcrProvider, SdkOcrProvider
from invoice_agent.pipeline import analyze_records, export_records, organize_batch_subfolders, organize_folder, resolve_trip_info
from invoice_agent.trip_audit import TripAuditPolicy, run_trip_audit
from invoice_agent.web import (
    TASKS,
    export_all_batch_packages,
    export_batch_package,
    export_task,
    get_task_snapshot,
    handle_choose_path,
    render_index,
    run_organize_from_form,
    update_task_record,
)


class FakeOcrProvider:
    def parse(self, path: Path) -> ParsedDocument:
        name = path.name
        if "trip" in name:
            return ParsedDocument(
                source_path=path,
                raw_text="滴滴出行行程报销单 总金额 88.00 起点 杭州 终点 上海",
                raw_result={"kind": "trip"},
                fields={
                    "document_type": "行程单",
                    "issue_date": "2026-03-02",
                    "total_amount": "88.00",
                    "origin": "杭州",
                    "destination": "上海",
                    "description": "杭州-上海",
                },
                ok=True,
            )
        if "duplicate" in name:
            invoice_number = "INV-DUP"
        else:
            invoice_number = "INV-001"
        return ParsedDocument(
            source_path=path,
            raw_text="增值税电子普通发票 发票号码 {} 价税合计 88.00".format(invoice_number),
            raw_result={"kind": "invoice"},
            fields={
                "document_type": "网约车发票",
                "issue_date": "2026-03-02",
                "invoice_number": invoice_number,
                "invoice_code": "",
                "seller_name": "滴滴出行科技有限公司",
                "buyer_name": "杭州勤合能源科技有限公司",
                "total_amount": "77.88",
                "total_tax": "10.12",
                "total_with_tax": "88.00",
                "description": "网约车服务",
            },
            ok=True,
        )


class SlowRecordingProvider(FakeOcrProvider):
    def __init__(self):
        self.active = 0
        self.max_active = 0
        self.lock = threading.Lock()

    def parse(self, path: Path) -> ParsedDocument:
        with self.lock:
            self.active += 1
            self.max_active = max(self.max_active, self.active)
        try:
            time.sleep(0.05)
            return super().parse(path)
        finally:
            with self.lock:
                self.active -= 1


class BlockingSecondPackageProvider(FakeOcrProvider):
    def __init__(self):
        self.started_second = threading.Event()
        self.release_second = threading.Event()

    def parse(self, path: Path) -> ParsedDocument:
        if "/B/" in str(path):
            self.started_second.set()
            self.release_second.wait(timeout=2)
        return super().parse(path)


class FakeResponse:
    def __init__(self, status_code=200, payload=None, text=""):
        self.status_code = status_code
        self._payload = payload or {}
        self.text = text

    def json(self):
        return self._payload

    def raise_for_status(self):
        if self.status_code >= 400:
            raise RuntimeError(self.text)


class FakeAsyncSession:
    def __init__(self):
        self.submitted = []
        self.poll_count = {}

    def post(self, url, headers=None, data=None, files=None, timeout=None):
        job_id = f"job-{len(self.submitted)}"
        self.submitted.append(files["file"].name)
        return FakeResponse(payload={"data": {"jobId": job_id}})

    def get(self, url, headers=None, timeout=None):
        if url.startswith("https://result/"):
            jsonl = json.dumps(
                {
                    "result": {
                        "layoutParsingResults": [
                            {
                                "markdown": {
                                    "text": "电子发票 发票号码：INVASYNC 开票日期：2026年03月01日 价税合计（小写）￥12.00"
                                }
                            }
                        ]
                    }
                },
                ensure_ascii=False,
            )
            return FakeResponse(text=jsonl)
        job_id = url.rsplit("/", 1)[-1]
        self.poll_count[job_id] = self.poll_count.get(job_id, 0) + 1
        return FakeResponse(
            payload={
                "data": {
                    "state": "done",
                    "extractProgress": {"extractedPages": 1},
                    "resultUrl": {"jsonUrl": f"https://result/{job_id}.jsonl"},
                }
            }
        )


def write_file(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def write_pdf(path: Path) -> None:
    fitz = pytest.importorskip("fitz")
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    page = doc.new_page(width=240, height=120)
    page.insert_text((32, 64), "invoice preview")
    doc.save(path)
    doc.close()


def make_record(
    tmp_path: Path,
    name: str,
    document_type: str,
    amount: str,
    seller_name: str = "",
    description: str = "",
    trip_start_date: str = "2026-03-01",
    trip_end_date: str = "2026-03-02",
    daily_meal_allowance: str = "0",
) -> ExpenseRecord:
    path = tmp_path / name
    write_file(path, name.encode("utf-8"))
    return ExpenseRecord(
        sequence=0,
        source_path=path,
        original_name=name,
        file_hash=name,
        project_name="项目",
        traveler="张三",
        department="技术部",
        trip_start_date=trip_start_date,
        trip_end_date=trip_end_date,
        daily_meal_allowance=daily_meal_allowance,
        document_date="2026-03-01",
        document_type=document_type,
        include_in_amount=True,
        invoice_number=name,
        seller_name=seller_name,
        total_with_tax=amount,
        description=description,
        recognition_status="已识别",
    )


def test_build_preview_includes_review_cards_and_lightweight_overview(tmp_path: Path):
    record = make_record(
        tmp_path,
        "invoice.pdf",
        "网约车发票",
        "88.00",
        seller_name="滴滴出行科技有限公司",
        description="杭州-上海",
    )
    record.sequence = 1
    record.high_level_category = "市区交通费"
    record.buyer_name = "杭州勤合能源科技有限公司"
    record.risk_note = "建议人工确认"

    preview = build_preview([record])

    assert preview["review_cards"] == [
        {
            "项目名称": "项目",
            "序号": 1,
            "原文件名": "invoice.pdf",
            "凭证日期": "2026-03-01",
            "凭证类别": "网约车发票",
            "报销大类": "市区交通费",
            "价税合计": 88.0,
            "是否计入金额": "是",
            "发票号码": "invoice.pdf",
            "销方名称": "滴滴出行科技有限公司",
            "购方名称": "杭州勤合能源科技有限公司",
            "起点": "",
            "终点": "",
            "行程/住宿说明": "杭州-上海",
            "风险提示": "建议人工确认",
        }
    ]
    assert preview["overview_rows"] == [
        {
            "序号": 1,
            "日期": "2026-03-01",
            "凭证类别": "网约车发票",
            "报销大类": "市区交通费",
            "金额": 88.0,
            "是否计入": "是",
            "风险": "建议人工确认",
            "原文件名": "invoice.pdf",
        }
    ]
    assert set(preview["overview_rows"][0]) == {
        "序号",
        "日期",
        "凭证类别",
        "报销大类",
        "金额",
        "是否计入",
        "风险",
        "原文件名",
    }
    assert preview["main_rows"][0]["原文件名"] == "invoice.pdf"


def test_task_preview_file_resolves_only_current_task_records(tmp_path: Path):
    pdf_record = make_record(tmp_path, "invoice.pdf", "网约车发票", "88.00")
    image_record = make_record(tmp_path, "receipt.png", "纸质拍照发票", "18.00")
    pdf_record.sequence = 1
    image_record.sequence = 2
    task_id = "task-file-preview"
    TASKS[task_id] = {"_records": [pdf_record, image_record]}

    path, content_type = web.get_task_preview_file(task_id, "1")
    assert path == pdf_record.source_path
    assert content_type == "application/pdf"

    path, content_type = web.get_task_preview_file(task_id, "2")
    assert path == image_record.source_path
    assert content_type == "image/png"

    with pytest.raises(ValueError, match="Task not found"):
        web.get_task_preview_file("missing-task", "1")
    with pytest.raises(ValueError, match="File not found"):
        web.get_task_preview_file(task_id, "99")
    with pytest.raises(ValueError, match="Invalid file sequence"):
        web.get_task_preview_file(task_id, "../invoice.pdf")


def test_task_preview_image_renders_pdf_first_page_and_images(tmp_path: Path):
    pdf_path = tmp_path / "invoice.pdf"
    png_path = tmp_path / "receipt.png"
    pdf_record = make_record(tmp_path, "invoice.pdf", "网约车发票", "88.00")
    image_record = make_record(tmp_path, "receipt.png", "纸质拍照发票", "18.00")
    write_pdf(pdf_path)
    write_file(png_path, b"\x89PNG\r\n\x1a\nfake")
    pdf_record.source_path = pdf_path
    image_record.source_path = png_path
    pdf_record.sequence = 1
    image_record.sequence = 2
    task_id = "task-image-preview"
    TASKS[task_id] = {"_records": [pdf_record, image_record]}

    content, content_type = web.get_task_preview_image(task_id, "1")
    assert content_type == "image/png"
    assert content.startswith(b"\x89PNG\r\n\x1a\n")

    content, content_type = web.get_task_preview_image(task_id, "2")
    assert content_type == "image/png"
    assert content == b"\x89PNG\r\n\x1a\nfake"


def test_task_preview_image_falls_back_to_macos_quicklook_when_fitz_is_missing(tmp_path: Path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    write_file(pdf_path, b"%PDF-1.7\n")
    record = make_record(tmp_path, "invoice.pdf", "网约车发票", "88.00")
    record.source_path = pdf_path
    record.sequence = 1
    TASKS["task-quicklook-preview"] = {"_records": [record]}

    def fail_fitz(path):
        raise ValueError("PDF preview rendering requires PyMuPDF")

    def fake_quicklook(cmd, check, stdout, stderr):
        assert cmd[:3] == ["qlmanage", "-t", "-s"]
        output_dir = Path(cmd[cmd.index("-o") + 1])
        write_file(output_dir / "invoice.pdf.png", b"\x89PNG\r\n\x1a\nquicklook")

    monkeypatch.setattr(web, "render_pdf_preview_with_fitz", fail_fitz)
    monkeypatch.setattr(web.subprocess, "run", fake_quicklook)

    content, content_type = web.get_task_preview_image("task-quicklook-preview", "1")

    assert content_type == "image/png"
    assert content == b"\x89PNG\r\n\x1a\nquicklook"


def test_organize_folder_generates_preview_without_copying(tmp_path: Path):
    src = tmp_path / "上海出差"
    write_file(src / "b" / "trip.pdf", b"trip")
    write_file(src / "a" / "invoice.pdf", b"invoice")
    write_file(src / "a" / "duplicate.pdf", b"invoice")
    write_file(src / "ignored.txt", b"ignore")
    trip_info = src / "trip_info.json"
    trip_info.write_text(
        json.dumps(
            {
                "traveler": "张三",
                "department": "技术部",
                "trip_start_date": "2026-03-01",
                "trip_end_date": "2026-03-03",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    result = organize_folder(
        folder=src,
        trip_info_path=trip_info,
        out_dir=tmp_path / "out",
        apply=False,
        ocr_provider=FakeOcrProvider(),
    )

    assert result.output_dir.exists()
    assert (result.output_dir / "00_报销清单.xlsx").exists()
    assert (result.output_dir / "raw_results.json").exists()
    assert (result.output_dir / "rename_plan.json").exists()
    assert not (result.output_dir / "01_已识别_重命名").exists()
    assert len(result.records) == 3
    assert [record.original_name for record in result.records] == [
        "invoice.pdf",
        "duplicate.pdf",
        "trip.pdf",
    ]
    assert result.records[0].include_in_amount is True
    assert result.records[-1].include_in_amount is False
    assert result.records[-1].linked_group == result.records[0].linked_group
    assert result.records[1].include_in_amount is False
    assert "Hash重复" in result.records[1].duplicate_mark
    assert "重复发票不计入汇总" in result.records[1].risk_note
    assert result.records[0].new_name.startswith("001_2026-03-02_网约车发票_")
    assert "发票号INV-001" in result.records[0].new_name
    assert "不计金额" in result.records[-1].new_name

    workbook = load_workbook(result.output_dir / "00_报销清单.xlsx", data_only=False)
    assert workbook.sheetnames == ["报销清单", "类别汇总", "公司报销表单汇总", "重复与风险", "行程校对", "重命名计划"]
    main = workbook["报销清单"]
    main_headers = [cell.value for cell in main[1]]
    assert "发票代码" not in main_headers
    assert main_headers[8] == "报销大类"
    main_rows = list(main.iter_rows(min_row=2, values_only=True))
    rows_by_name = {row[19]: row for row in main_rows if row[0] != "合计总金额"}
    assert rows_by_name["invoice.pdf"][8] == "交通费"
    assert rows_by_name["duplicate.pdf"][8] in {None, ""}
    assert rows_by_name["trip.pdf"][8] in {None, ""}
    category_total = sum(row[15] for row in main_rows if row[9] == "是" and row[8] == "交通费")
    assert category_total == 88.0
    assert main.cell(row=main.max_row, column=1).value == "合计总金额"
    assert float(main.cell(row=main.max_row, column=16).value) == 88.0
    summary = workbook["类别汇总"]
    summary_rows = {row[0]: (float(row[1]), row[2]) for row in summary.iter_rows(min_row=2, values_only=True)}
    assert summary["A1"].value == "报销大类"
    assert list(summary_rows) == [
        "行程交通费",
        "住宿费",
        "市区交通费",
        "通行费",
        "过路费",
        "油费",
        "退改费",
        "其他费用",
        "出差餐补",
        "合计总金额",
    ]
    assert summary_rows["通行费"] == (88.0, 1)
    assert summary_rows["行程交通费"] == (0.0, 0)
    assert summary_rows["出差餐补"] == (150.0, 0)
    assert summary_rows["合计总金额"] == (238.0, 1)


def test_batch_mode_discovers_first_level_subfolders_only(tmp_path: Path):
    root = tmp_path / "待整理报销"
    write_file(root / "A" / "nested" / "invoice.pdf", b"invoice-a")
    write_file(root / "B" / "invoice.pdf", b"invoice-b")
    write_file(root / ".hidden" / "invoice.pdf", b"hidden")
    write_file(root / "整理结果" / "invoice.pdf", b"output")

    result = organize_batch_subfolders(
        folder=root,
        trip_info=TripInfo("公共", "张三", "技术部", "2026-03-01", "2026-03-02"),
        out_dir=root / "整理结果",
        ocr_provider=FakeOcrProvider(),
        write_excel=False,
    )

    assert [item.name for item in result.items] == ["A", "B"]
    assert result.items[0].state == "review"
    assert [record.original_name for record in result.items[0].result.records] == ["invoice.pdf"]


def test_batch_mode_fails_when_root_contains_direct_invoice_files(tmp_path: Path):
    root = tmp_path / "待整理报销"
    write_file(root / "invoice.pdf", b"direct")
    write_file(root / "A" / "invoice.pdf", b"nested")

    with pytest.raises(ValueError, match="直属票据"):
        organize_batch_subfolders(
            folder=root,
            trip_info=TripInfo("公共", "张三", "技术部", "2026-03-01", "2026-03-02"),
            ocr_provider=FakeOcrProvider(),
        )


def test_batch_mode_processes_each_subfolder_as_independent_package(tmp_path: Path):
    root = tmp_path / "待整理报销"
    for name in ["A", "B"]:
        write_file(root / name / "invoice.pdf", f"invoice-{name}".encode("utf-8"))
        (root / name / "trip_info.json").write_text(
            json.dumps(
                {
                    "traveler": name,
                    "department": "技术部",
                    "trip_start_date": "2026-03-01",
                    "trip_end_date": "2026-03-02",
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    result = organize_batch_subfolders(root, out_dir=tmp_path / "out", ocr_provider=FakeOcrProvider())

    assert [item.state for item in result.items] == ["done", "done"]
    a_raw = json.loads((tmp_path / "out" / "A" / "raw_results.json").read_text(encoding="utf-8"))
    b_raw = json.loads((tmp_path / "out" / "B" / "raw_results.json").read_text(encoding="utf-8"))
    assert [row["source_path"] for row in a_raw] == [str((root / "A" / "invoice.pdf").resolve())]
    assert [row["source_path"] for row in b_raw] == [str((root / "B" / "invoice.pdf").resolve())]
    assert (tmp_path / "out" / "A" / "00_报销清单.xlsx").exists()
    assert (tmp_path / "out" / "B" / "00_报销清单.xlsx").exists()


def test_batch_mode_continues_when_one_package_missing_trip_info(tmp_path: Path):
    root = tmp_path / "待整理报销"
    write_file(root / "A" / "invoice.pdf", b"invoice-a")
    write_file(root / "B" / "invoice.pdf", b"invoice-b")
    (root / "B" / "trip_info.json").write_text(
        json.dumps(
            {
                "traveler": "李四",
                "department": "技术部",
                "trip_start_date": "2026-03-01",
                "trip_end_date": "2026-03-02",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    result = organize_batch_subfolders(root, out_dir=tmp_path / "out", ocr_provider=FakeOcrProvider())

    states = {item.name: item.state for item in result.items}
    assert states == {"A": "failed", "B": "done"}
    assert "Missing trip info" in next(item.error for item in result.items if item.name == "A")
    assert (tmp_path / "out" / "B" / "00_报销清单.xlsx").exists()


def test_category_summary_uses_fixed_reimbursement_categories(tmp_path: Path):
    records = [
        make_record(tmp_path, "train.pdf", "高铁发票", "10.00"),
        make_record(tmp_path, "hotel.pdf", "住宿发票", "20.00"),
        make_record(tmp_path, "didi.pdf", "网约车发票", "30.00"),
        make_record(tmp_path, "toll.pdf", "通行费发票", "40.00"),
        make_record(tmp_path, "fuel.pdf", "普票", "50.00", seller_name="中国石化销售有限公司"),
        make_record(tmp_path, "refund.pdf", "高铁发票", "60.00", description="退票费"),
        make_record(
            tmp_path,
            "meal.pdf",
            "餐饮发票",
            "70.00",
            seller_name="饭店",
            trip_start_date="2026-01-01",
            trip_end_date="2026-01-05",
            daily_meal_allowance="100",
        ),
    ]

    analyze_records(records)
    write_workbook(tmp_path / "summary.xlsx", records)

    workbook = load_workbook(tmp_path / "summary.xlsx", data_only=True)
    rows = {row[0]: (float(row[1]), row[2]) for row in workbook["类别汇总"].iter_rows(min_row=2, values_only=True)}
    assert rows == {
        "行程交通费": (10.0, 1),
        "住宿费": (20.0, 1),
        "市区交通费": (0.0, 0),
        "通行费": (30.0, 1),
        "过路费": (40.0, 1),
        "油费": (50.0, 1),
        "退改费": (60.0, 1),
        "其他费用": (70.0, 1),
        "出差餐补": (500.0, 0),
        "合计总金额": (780.0, 7),
    }

    preview_rows = {
        row["报销大类"]: (row["计入金额合计"], row["张数"])
        for row in build_preview(records)["summary_rows"]
    }
    assert preview_rows["出差餐补"] == (500.0, 0)
    assert preview_rows["合计总金额"] == (780.0, 7)


def test_chinese_currency_uppercase():
    assert amount_to_chinese_upper(Decimal("6157.11")) == "人民币陆仟壹佰伍拾柒元壹角壹分"
    assert amount_to_chinese_upper(Decimal("100.00")) == "人民币壹佰元整"
    assert amount_to_chinese_upper(Decimal("0.05")) == "人民币伍分"


def test_build_company_data_filters_itineraries_and_maps_categories(tmp_path: Path):
    train = make_record(tmp_path, "train.pdf", "高铁发票", "232.00", seller_name="铁路电子客票", description="杭州东-蚌埠南")
    train.origin = "杭州东"
    train.destination = "蚌埠南"
    ride = make_record(tmp_path, "ride.pdf", "网约车发票", "88.00", seller_name="滴滴出行科技有限公司", description="酒店-项目现场")
    hotel = make_record(
        tmp_path,
        "hotel.pdf",
        "住宿发票",
        "2380.00",
        seller_name="维纳民宿酒店",
        daily_meal_allowance="50",
    )
    material = make_record(
        tmp_path,
        "material.pdf",
        "普票",
        "233.00",
        seller_name="顺安消防器材经营部",
        description="购买灭火器材料",
    )
    itinerary = make_record(tmp_path, "trip.pdf", "行程单", "88.00", description="酒店-项目现场")
    itinerary.include_in_amount = False
    records = [train, ride, hotel, material, itinerary]
    analyze_records(records)

    data = build_company_reimbursement_data(records, export_date=date(2026, 6, 22))

    assert [row.kind for row in data.travel_rows] == ["intercity", "city"]
    assert data.lodging_total == Decimal("2380.00")
    assert data.daily_rows[0].content == "购买灭火器材料"
    assert data.meal_allowance.amount == Decimal("100.00")
    assert data.meal_allowance.count == 0
    assert data.total == Decimal("3033.00")


def test_company_template_has_clean_required_sheets():
    workbook = load_workbook(company_template_path(), data_only=False)

    assert workbook.sheetnames == ["报销明细表", "差旅费报销单", "日常费用报销单"]
    assert str(workbook["报销明细表"].page_setup.paperSize) == "9"
    assert workbook["报销明细表"].page_setup.orientation == "portrait"
    assert str(workbook["差旅费报销单"].page_setup.paperSize) == "11"
    assert workbook["差旅费报销单"].page_setup.orientation == "landscape"
    assert str(workbook["日常费用报销单"].page_setup.paperSize) == "11"
    assert workbook["日常费用报销单"].page_setup.orientation == "landscape"
    assert workbook["报销明细表"]["B2"].value in {None, ""}
    assert workbook["差旅费报销单"]["E3"].value in {None, ""}
    assert workbook["日常费用报销单"]["F3"].value in {None, ""}


def test_company_workbook_omits_empty_daily_sheet(tmp_path: Path):
    train = make_record(tmp_path, "train-only.pdf", "高铁发票", "232.00", description="杭州东-蚌埠南")
    analyze_records([train])
    data = build_company_reimbursement_data([train], export_date=date(2026, 6, 22))

    output = tmp_path / "company.xlsx"
    write_company_workbook(output, data)
    workbook = load_workbook(output, data_only=False)

    assert workbook.sheetnames == ["报销明细表", "差旅费报销单"]
    assert workbook["报销明细表"]["B2"].value == "项目"
    assert workbook["差旅费报销单"]["E3"].value == "张三"
    detail_labels = [workbook["报销明细表"].cell(row, 1).value for row in range(1, workbook["报销明细表"].max_row + 1)]
    assert detail_labels.count("总计") == 1
    assert detail_labels.count("报销人：") == 1


def test_company_workbook_paginates_daily_and_travel_rows(tmp_path: Path):
    daily_records = [
        make_record(tmp_path, f"material-{index}.pdf", "普票", "10.00", description=f"材料 {index}")
        for index in range(9)
    ]
    travel_records = [
        make_record(tmp_path, f"train-{index}.pdf", "高铁发票", "20.00", description=f"行程 {index}")
        for index in range(13)
    ]
    for record in travel_records:
        record.origin = "杭州"
        record.destination = "上海"
    records = daily_records + travel_records
    analyze_records(records)
    data = build_company_reimbursement_data(records, export_date=date(2026, 6, 22))

    output = tmp_path / "company-pages.xlsx"
    write_company_workbook(output, data)
    workbook = load_workbook(output, data_only=False)

    assert [name for name in workbook.sheetnames if name.startswith("日常费用报销单")] == [
        "日常费用报销单",
        "日常费用报销单 (2)",
    ]
    assert [name for name in workbook.sheetnames if name.startswith("差旅费报销单")] == [
        "差旅费报销单",
        "差旅费报销单 (2)",
    ]
    assert workbook["日常费用报销单 (2)"]["B5"].value == "材料 8"
    assert workbook["差旅费报销单 (2)"]["B6"].value == "2026-03-01"


def test_pdf_export_skips_when_excel_unavailable(monkeypatch, tmp_path: Path):
    monkeypatch.setattr(pdf_export, "find_excel_executable", lambda: None)

    result = pdf_export.export_company_pdf(tmp_path / "company.xlsx", tmp_path / "company.pdf")

    assert result.status == "skipped"
    assert "Microsoft Excel" in result.message


def test_pdf_export_reports_subprocess_failure(monkeypatch, tmp_path: Path):
    monkeypatch.setattr(pdf_export, "find_excel_executable", lambda: Path("EXCEL.EXE"))
    monkeypatch.setattr(
        pdf_export.subprocess,
        "run",
        lambda *args, **kwargs: type("Result", (), {"returncode": 1, "stdout": "", "stderr": "COM failed"})(),
    )

    result = pdf_export.export_company_pdf(tmp_path / "company.xlsx", tmp_path / "company.pdf")

    assert result.status == "failed"
    assert "COM failed" in result.message


def test_export_records_generates_company_excel_and_preserves_summary(tmp_path: Path, monkeypatch):
    record = make_record(tmp_path, "invoice-export.pdf", "网约车发票", "88.00", seller_name="滴滴出行科技有限公司")
    analyze_records([record])
    monkeypatch.setattr(
        "invoice_agent.pipeline.export_company_pdf",
        lambda source, target: pdf_export.PdfExportResult("skipped", target, "test skip"),
    )

    result = export_records(tmp_path / "out", [record], write_excel=True)

    assert (tmp_path / "out" / "00_报销清单.xlsx").exists()
    assert (tmp_path / "out" / "01_公司报销单.xlsx").exists()
    assert result.status == "success"
    assert {artifact.kind for artifact in result.artifacts} == {"summary_excel", "company_excel", "company_pdf"}


def test_export_records_keeps_summary_when_company_export_fails(tmp_path: Path, monkeypatch):
    record = make_record(tmp_path, "invoice-fail.pdf", "网约车发票", "88.00")
    analyze_records([record])
    monkeypatch.setattr(
        "invoice_agent.pipeline.write_company_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("template broken")),
    )

    result = export_records(tmp_path / "out", [record], write_excel=True)

    assert (tmp_path / "out" / "00_报销清单.xlsx").exists()
    assert not (tmp_path / "out" / "01_公司报销单.xlsx").exists()
    assert result.status == "partial_success"
    assert "template broken" in result.warnings[0]


def test_workbook_and_preview_include_trip_audit_rows(tmp_path: Path):
    outbound = make_record(tmp_path, "train-out.pdf", "高铁发票", "130.00", trip_start_date="2026-03-01", trip_end_date="2026-03-03")
    outbound.sequence = 1
    outbound.document_date = "2026-03-01"
    outbound.origin = "杭州东"
    outbound.destination = "南京南"
    hotel = make_record(tmp_path, "hotel.pdf", "住宿发票", "900.00", description="住宿 1晚", trip_start_date="2026-03-01", trip_end_date="2026-03-03")
    hotel.sequence = 2
    audit = run_trip_audit([outbound, hotel], TripAuditPolicy(lodging_daily_limit="300"))

    preview = build_preview([outbound, hotel], audit)
    write_workbook(tmp_path / "audit.xlsx", [outbound, hotel], audit)

    assert preview["trip_audit_rows"][0]["校对类别"]
    workbook = load_workbook(tmp_path / "audit.xlsx", data_only=True)
    assert "行程校对" in workbook.sheetnames
    rows = list(workbook["行程校对"].iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "住宿标准" for row in rows)


def test_trip_audit_detects_missing_return_lodging_mismatch_and_city_transport_over_limit(tmp_path: Path):
    outbound = make_record(tmp_path, "train-out.pdf", "高铁发票", "130.00", trip_start_date="2026-03-01", trip_end_date="2026-03-05")
    outbound.document_date = "2026-03-01"
    outbound.origin = "杭州东"
    outbound.destination = "南京南"
    hotel = make_record(tmp_path, "hotel.pdf", "住宿发票", "1000.00", description="酒店住宿 3晚", trip_start_date="2026-03-01", trip_end_date="2026-03-05")
    hotel.document_date = "2026-03-02"
    city_ride_1 = make_record(tmp_path, "didi-1.pdf", "网约车发票", "80.00", seller_name="滴滴出行科技有限公司", trip_start_date="2026-03-01", trip_end_date="2026-03-05")
    city_ride_1.document_date = "2026-03-02"
    city_ride_2 = make_record(tmp_path, "didi-2.pdf", "出租车票", "50.00", trip_start_date="2026-03-01", trip_end_date="2026-03-05")
    city_ride_2.document_date = "2026-03-02"
    records = [outbound, hotel, city_ride_1, city_ride_2]
    for index, record in enumerate(records, start=1):
        record.sequence = index

    result = run_trip_audit(records, TripAuditPolicy(city_transport_daily_limit="100", lodging_daily_limit="200"))

    conclusions = {item.category: item.conclusion for item in result.items}
    assert "未发现出差结束日期对应的返程城际交通票" in conclusions["日期覆盖"]
    assert "最后一段城际交通未返回杭州" in conclusions["行程闭环"]
    assert "住宿发票晚数 3 晚，少于预计 4 晚" in conclusions["住宿晚数"]
    assert "2026-03-02 市内交通 130.00 元，超过每日标准 100.00 元" in conclusions["市内交通"]
    assert "住宿费 1000.00 元，超过 4 晚标准 800.00 元" in conclusions["住宿标准"]
    assert result.items[0].evidence


def test_trip_audit_llm_uses_structured_payload_without_raw_text_and_falls_back_on_error(tmp_path: Path):
    record = make_record(tmp_path, "train-out.pdf", "高铁发票", "130.00", trip_start_date="2026-03-01", trip_end_date="2026-03-02")
    record.sequence = 1
    record.document_date = "2026-03-01"
    record.origin = "杭州东"
    record.destination = "南京南"
    record.raw_text = "SHOULD_NOT_BE_SENT_TO_LLM"
    captured = {}

    class FailingClient:
        def review(self, payload):
            captured["payload"] = payload
            raise RuntimeError("llm unavailable")

    result = run_trip_audit([record], TripAuditPolicy(enable_llm_review=True), llm_client=FailingClient())

    assert captured["payload"]["records"][0]["sequence"] == 1
    assert "raw_text" not in captured["payload"]["records"][0]
    assert "SHOULD_NOT_BE_SENT_TO_LLM" not in json.dumps(captured["payload"], ensure_ascii=False)
    assert any(item.category == "模型复核" and "模型复核未执行" in item.conclusion for item in result.items)


def test_trip_info_defaults_daily_meal_allowance_to_50(tmp_path: Path):
    trip_info_path = tmp_path / "trip_info.json"
    trip_info_path.write_text(
        json.dumps(
            {
                "project_name": "上海出差",
                "traveler": "张三",
                "department": "技术部",
                "trip_start_date": "2026-01-01",
                "trip_end_date": "2026-01-05",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    trip_info = resolve_trip_info(tmp_path, trip_info_path=trip_info_path)

    assert trip_info.daily_meal_allowance == "50"


def test_apply_copies_files_and_preserves_originals(tmp_path: Path):
    src = tmp_path / "杭州出差"
    invoice = src / "invoice.pdf"
    write_file(invoice, b"invoice")
    before = invoice.read_bytes()

    result = organize_folder(
        folder=src,
        trip_info=TripInfo(
            project_name="杭州出差",
            traveler="李四",
            department="财务部",
            trip_start_date="2026-03-01",
            trip_end_date="2026-03-02",
        ),
        out_dir=tmp_path / "out",
        apply=True,
        ocr_provider=FakeOcrProvider(),
    )

    copied = result.output_dir / "01_已识别_重命名" / result.records[0].new_name
    assert copied.exists()
    assert copied.read_bytes() == before
    assert invoice.read_bytes() == before


def test_organize_folder_can_parse_documents_concurrently(tmp_path: Path):
    src = tmp_path / "并发测试"
    for index in range(4):
        write_file(src / f"invoice-{index}.pdf", f"invoice-{index}".encode("utf-8"))
    provider = SlowRecordingProvider()

    organize_folder(
        folder=src,
        trip_info=TripInfo(
            project_name="并发测试",
            traveler="张三",
            department="技术部",
            trip_start_date="2026-03-01",
            trip_end_date="2026-03-02",
        ),
        out_dir=tmp_path / "out",
        apply=False,
        ocr_provider=provider,
        max_workers=3,
    )

    assert provider.max_active > 1


def test_load_agent_config_reads_local_paddleocr_settings(tmp_path: Path):
    config_path = tmp_path / "invoice_agent_config.json"
    config_path.write_text(
        json.dumps(
            {
                "paddleocr_doc_parsing_api_url": "https://example.com/layout-parsing",
                "paddleocr_job_url": "https://example.com/jobs",
                "paddleocr_access_token": "token-from-file",
                "paddleocr_model": "PaddleOCR-VL-1.6",
                "ocr_provider": "async_jobs",
                "city_transport_daily_limit": "120",
                "lodging_daily_limit": "450",
                "llm_base_url": "https://llm.example.com/v1",
                "llm_model": "compatible-model",
                "llm_api_key_env": "INVOICE_AGENT_LLM_API_KEY",
            }
        ),
        encoding="utf-8",
    )

    config = load_agent_config(config_path)

    assert config.paddleocr_doc_parsing_api_url == "https://example.com/layout-parsing"
    assert config.paddleocr_job_url == "https://example.com/jobs"
    assert config.paddleocr_access_token == "token-from-file"
    assert config.ocr_provider == "async_jobs"  # 自定义配置值
    assert config.city_transport_daily_limit == "120"
    assert config.lodging_daily_limit == "450"
    assert config.llm_base_url == "https://llm.example.com/v1"
    assert config.llm_model == "compatible-model"
    assert config.llm_api_key_env == "INVOICE_AGENT_LLM_API_KEY"


def test_async_paddle_provider_submits_all_jobs_before_polling(tmp_path: Path):
    paths = []
    for index in range(2):
        path = tmp_path / f"invoice-{index}.pdf"
        write_file(path, b"pdf")
        paths.append(path)
    session = FakeAsyncSession()
    provider = PaddleAsyncOcrProvider(
        job_url="https://example.com/jobs",
        access_token="token",
        poll_interval_seconds=0,
        timeout_seconds=10,
        session=session,
    )

    docs = provider.parse_many(paths, max_workers=2)

    assert len(session.submitted) == 2
    assert [doc.ok for doc in docs] == [True, True]
    assert docs[0].fields["invoice_number"] == "INVASYNC"


def test_sdk_provider_returns_error_without_token(tmp_path: Path):
    path = tmp_path / "invoice.pdf"
    write_file(path, b"pdf")
    provider = SdkOcrProvider(access_token="")

    docs = provider.parse_many([path])

    assert len(docs) == 1
    assert docs[0].ok is False
    assert docs[0].error["code"] == "CONFIG_ERROR"


def test_sdk_provider_catches_sdk_errors_gracefully(tmp_path: Path, monkeypatch):
    path = tmp_path / "invoice.pdf"
    write_file(path, b"pdf")
    provider = SdkOcrProvider(access_token="invalid-token", timeout_seconds=1)

    docs = provider.parse_many([path])

    assert len(docs) == 1
    assert docs[0].ok is False


def test_extracts_train_ticket_amount_and_travel_date():
    fields = extract_fields_from_text(
        "电子发票（铁路电子客票） 发票号码：26329116804002389322 "
        "开票日期：2026年03月13日 南京南站 G7629 杭州东站 "
        "2026年03月06日 19:54开 票价：￥130.00",
        Path("2026-03-06 火车票-江苏南京南→杭州东 G7629 ¥130.pdf"),
    )

    assert fields["document_type"] == "高铁发票"
    assert fields["issue_date"] == "2026-03-06"
    assert fields["total_with_tax"] == "130.00"


def test_extracts_train_parties_and_route_from_ocr_text():
    fields = extract_fields_from_text(
        "# 电子发票（铁路电子客票）\n\n"
        "发票号码：26329116804002389321\n\n"
        "开票日期：2026年03月13日\n\n"
        "南京南站\n\n"
        "Nanjingnan\n\n"
        "G736\n\n"
        "蚌埠南站\n\n"
        "2026年02月04日 12:45开\n\n"
        "Bengbunan\n\n"
        "票价：￥87.00\n\n"
        "购买方名称：杭州勤合能源科技有限公司\n\n"
        "统一社会信用代码：91330105MA7C4PL8XP\n\n",
        Path("2.pdf"),
    )

    assert fields["document_type"] == "高铁发票"
    assert fields["seller_name"] == "铁路电子客票"
    assert fields["buyer_name"] == "杭州勤合能源科技有限公司"
    assert fields["origin"] == "南京南站"
    assert fields["destination"] == "蚌埠南站"
    assert fields["description"] == "南京南站-蚌埠南站"


def test_extracts_didi_itinerary_route_from_table():
    fields = extract_fields_from_text(
        "滴滴出行-行程单 共1笔行程，合计17.50元"
        "<table><tr><td>序号</td><td>车型</td><td>上车时间</td><td>城市</td><td>起点</td><td>终点</td><td>里程[公里]</td><td>金额[元]</td></tr>"
        "<tr><td>1</td><td>特惠快车</td><td>02-04 09:35 周三</td><td>南京市</td><td>东屏|溧水站-站前广场</td><td>溧水区|国家电网(溧水城区供电营业厅)</td><td>9.5</td><td>17.50</td></tr></table>",
        Path("滴滴出行行程报销单.pdf"),
    )

    assert fields["document_type"] == "行程单"
    assert fields["origin"] == "东屏|溧水站-站前广场"
    assert fields["destination"] == "溧水区|国家电网(溧水城区供电营业厅)"
    assert fields["description"] == "东屏|溧水站-站前广场-溧水区|国家电网(溧水城区供电营业厅)"


def test_hotel_invoice_with_tax_id_is_not_classified_as_train():
    fields = extract_fields_from_text(
        "电子发票（增值税专用发票） 发票号码：26342000000740203651 "
        "开票日期：2026年03月12日 购买方名称: 杭州勤合能源科技有限公司 "
        "统一社会信用代码：91330105MA7C4PL8XP 项目名称 *住宿服务*住宿费 "
        "价税合计（小写）￥183.92",
        Path("2026-03-12 住宿服务-蚌埠市十全十美酒店管理有限公司.pdf"),
    )

    assert fields["document_type"] == "住宿发票"
    assert fields["total_with_tax"] == "183.92"


def test_extracts_itinerary_total_amount():
    didi = extract_fields_from_text(
        "滴滴出行-行程单 共4笔行程，合计74.00元",
        Path("滴滴出行行程报销单A.pdf"),
    )
    toll = extract_fields_from_text(
        "江苏省车辆通行费电子票据行程单 累计金额（元） 110.00 行程数量 1",
        Path("江苏高速通行费电子票据行程单_20260322.pdf"),
    )

    assert didi["document_type"] == "行程单"
    assert didi["total_with_tax"] == "74.00"
    assert toll["document_type"] == "行程单"
    assert toll["total_with_tax"] == "110.00"


def test_route_extraction_ignores_dates_and_plain_hyphens():
    meal = extract_fields_from_text(
        "电子发票 发票号码：26344000000026702641 开票日期：2026年02月24日",
        Path("2026-02-24 餐饮服务-蚌埠市蚌山区沁恬饮品店 ¥50.pdf"),
    )
    toll = extract_fields_from_text(
        "电子发票 发票号码：26347000000063949395 开票日期：2026年03月22日",
        Path("【浙AGP2113】通行发票-1.pdf"),
    )

    assert meal["origin"] == ""
    assert meal["destination"] == ""
    assert toll["origin"] == ""
    assert toll["destination"] == ""


def test_route_extraction_uses_arrow_route_from_train_filename():
    fields = extract_fields_from_text(
        "电子发票（铁路电子客票） 发票号码：26329116804002389322 "
        "开票日期：2026年03月13日 2026年03月06日 19:54开 票价：￥130.00",
        Path("2026-03-06 火车票-江苏南京南→杭州东 G7629 ¥130.pdf"),
    )

    assert fields["origin"] == "江苏南京南"
    assert fields["destination"] == "杭州东"
    assert fields["description"] == "江苏南京南-杭州东"


def test_cli_supports_ui_command():
    parser = build_parser()
    args = parser.parse_args(["ui", "--host", "127.0.0.1", "--port", "9999"])

    assert args.command == "ui"
    assert args.host == "127.0.0.1"
    assert args.port == 9999


def test_cli_organize_supports_speed_options():
    parser = build_parser()
    args = parser.parse_args(
        [
            "organize",
            "测试发票",
            "--max-workers",
            "4",
            "--timeout-seconds",
            "60",
            "--ocr-provider",
            "async_jobs",
            "--city-transport-daily-limit",
            "120",
            "--lodging-daily-limit",
            "450",
            "--enable-llm-review",
        ]
    )

    assert args.max_workers == 4
    assert args.timeout_seconds == 60
    assert args.ocr_provider == "async_jobs"
    assert args.city_transport_daily_limit == "120"
    assert args.lodging_daily_limit == "450"
    assert args.enable_llm_review is True


def test_cli_accepts_batch_subfolders_mode():
    parser = build_parser()
    args = parser.parse_args(["organize", "待整理报销", "--mode", "batch-subfolders"])

    assert args.mode == "batch-subfolders"


def test_cli_trip_info_defaults_daily_meal_allowance():
    parser = build_parser()
    args = parser.parse_args(
        [
            "organize",
            "测试发票",
            "--traveler",
            "张三",
            "--department",
            "技术部",
            "--trip-start-date",
            "2026-01-01",
            "--trip-end-date",
            "2026-01-05",
        ]
    )

    trip_info = _trip_info_from_args(args)

    assert trip_info is not None
    assert trip_info.daily_meal_allowance == "50"


def test_frontend_renders_pdf_previews_with_image_lightbox():
    js = Path("invoice_agent/static/app.js").read_text(encoding="utf-8")

    assert "/previews/" in js
    assert "/files/" in js
    assert "/records/" in js
    assert "renderFilePreview(previewUrl, fileUrl, fileName, isPdf, isImage)" in js
    assert 'data-preview-url="' in js
    assert 'data-file-url="' in js
    assert 'data-preview-kind="image"' in js
    assert "previewButton.dataset.previewUrl," in js
    assert "kind === 'pdf'" not in js
    assert "<iframe" not in js
    assert "handlePreviewImageError" in js
    assert "preview-fallback-actions" in js
    assert "data-review-edit-toggle" in js
    assert "data-review-edit-cancel" in js
    assert "review-readonly" in js
    assert "review-summary-strip" in js
    assert "review-extra-details" in js
    assert "review-edit-extra" in js
    assert "data-review-edit" in js
    assert "saveReviewEdits" in js


def test_ui_page_contains_required_form_fields():
    page = render_index()
    css = Path("invoice_agent/static/app.css").read_text(encoding="utf-8")
    js = Path("invoice_agent/static/app.js").read_text(encoding="utf-8")

    for field in [
        '<link rel="stylesheet" href="/static/app.css?v=review-edit-v3">',
        '<script src="/static/app.js?v=review-edit-v3" defer></script>',
        'data-initial-task-id=""',
        'data-task-state="idle"',
        'class="app-shell"',
        'class="app-topbar"',
        'class="layout-shell"',
        'class="workspace-grid"',
        'class="workflow-stepper"',
        'data-flow-step="upload"',
        'data-flow-step="ocr"',
        'data-flow-step="review"',
        'data-flow-step="export"',
        "任务配置",
        "执行状态",
        "预览确认",
        "导出Excel",
        'name="folder"',
        'name="organize_mode"',
        'value="batch_subfolders"',
        'name="config"',
        'name="out_dir"',
        'name="traveler"',
        'name="department"',
        'name="trip_start_date"',
        'name="trip_end_date"',
        'name="daily_meal_allowance" type="text" value="50"',
        'name="city_transport_daily_limit" type="text" value="100"',
        'name="lodging_daily_limit"',
        'name="enable_llm_review"',
        'name="ocr_provider"',
        'name="max_workers"',
        'name="timeout_seconds"',
        'name="apply"',
        "choosePath('folder', 'folder')",
        "choosePath('config', 'file')",
        "choosePath('out_dir', 'folder')",
        'id="progress-status"',
        "报销清单",
        "图文核对",
        "表格总览",
        "类别汇总",
        "重复与风险",
        "行程校对",
        "重命名计划",
        "确认导出",
        "单报销包模式",
        "批量报销包模式",
        "高级 OCR 配置",
        'class="preview-block collapsible-preview-block"',
        "<summary>重命名计划</summary>",
    ]:
        assert field in page

    for field in [
        "任务已提交，正在准备",
        "/tasks/",
        "fetch('/organize'",
        "'X-Requested-With': 'fetch'",
        "new URLSearchParams(new FormData(form))",
        "initialTaskId",
        "setTaskState",
        "renderStepper",
        "data-flow-step",
        "报销大类",
        "trip_audit_rows",
        "renderTripAudit(preview || {})",
        "renderTripAuditCalendar",
        "renderTripAuditRiskPanel",
        "deriveTripAuditModel",
        "校对明细",
        "未归档风险",
        "导出中...",
        "已导出",
        "column-control-panel",
        "data-column-toggle",
        "toggleColumnVisibility",
        "data-table-key",
        "preview-table main-preview-table",
        "preview-table rename-preview-table",
        "renderReviewCards",
        "partial_success",
        "company_excel_path",
        "company_pdf_path",
        "renderBatchPackages",
        "/export-all",
        "/packages/",
        "renderOverviewTable",
        "openPreviewLightbox",
        "saveReviewEdits",
        "review-card",
        "/previews/",
    ]:
        assert field in js

    for field in [
        "--surface: #faf6f0",
        ".app-topbar",
        ".layout-shell",
        ".workflow-stepper",
        ".flow-step.is-current",
        ".flow-step div > span",
        ".flow-dot",
        "line-height: 1",
        ".column-control-panel",
        ".collapsible-preview-block",
        "grid-template-columns: minmax(340px, 390px) minmax(0, 1fr)",
        ".sticky-export",
        ".preview-table",
        ".review-card",
        ".review-card-preview",
        ".review-summary-strip",
        ".review-extra-details",
        ".trip-audit-calendar",
        ".trip-audit-risk-panel",
        ".trip-audit-detail-table",
        ".trip-audit-day.is-warning",
        ".review-readonly",
        ".review-edit-form",
        ".review-save-button",
        ".preview-lightbox",
        "word-break: keep-all",
        "min-width: 1760px",
        "@media (max-width: 860px)",
    ]:
        assert field in css


def test_web_flow_previews_before_exporting_excel(tmp_path: Path, monkeypatch):
    src = tmp_path / "上海出差"
    write_file(src / "invoice.pdf", b"invoice")
    out_dir = tmp_path / "out"
    task_id = "task-preview"
    TASKS[task_id] = {
        "id": task_id,
        "state": "queued",
        "stage": "排队中",
        "started_at": time.time(),
        "updated_at": time.time(),
        "total": 0,
        "completed": 0,
        "files": [],
        "output_dir": "",
        "excel_path": "",
        "error": "",
    }
    form = {
        "folder": str(src),
        "out_dir": str(out_dir),
        "traveler": "张三",
        "department": "技术部",
        "trip_start_date": "2026-03-01",
        "trip_end_date": "2026-03-02",
        "max_workers": "1",
        "timeout_seconds": "120",
    }
    monkeypatch.setattr(
        "invoice_agent.web.load_agent_config",
        lambda path: type(
            "Config",
            (),
            {
                "ocr_provider": "async_jobs",
                "paddleocr_job_url": "https://example.com/jobs",
                "paddleocr_access_token": "",
                "paddleocr_model": "PaddleOCR-VL-1.6",
                "request_timeout_seconds": 60,
                "retry_max_attempts": 3,
                "retry_base_delay_seconds": 1.0,
                "fallback_api_url": "",
            },
        )(),
    )
    monkeypatch.setattr("invoice_agent.web.SdkOcrProvider", lambda **kwargs: FakeOcrProvider())

    run_organize_from_form(form, task_id=task_id)

    task = TASKS[task_id]
    assert task["state"] == "review"
    assert task["can_export"] is True
    assert task["excel_path"] == ""
    assert not (out_dir / "00_报销清单.xlsx").exists()
    assert task["preview"]["main_rows"][0]["原文件名"] == "invoice.pdf"
    assert task["preview"]["main_rows"][0]["报销大类"] == "交通费"
    assert task["preview"]["summary_rows"][3]["报销大类"] == "通行费"
    assert task["preview"]["summary_rows"][-1]["报销大类"] == "合计总金额"
    assert task["preview"]["trip_audit_rows"]
    assert task["preview"]["main_rows"][-1]["序号"] == "合计总金额"
    assert task["preview"]["rename_rows"][0]["新文件名"].endswith(".pdf")

    export_result = export_task(task_id)

    assert export_result["excel_path"] == str(out_dir / "00_报销清单.xlsx")
    assert export_result["company_excel_path"] == str(out_dir / "01_公司报销单.xlsx")
    assert "export_artifacts" in export_result
    assert (out_dir / "00_报销清单.xlsx").exists()
    assert (out_dir / "01_公司报销单.xlsx").exists()
    workbook = load_workbook(out_dir / "00_报销清单.xlsx", data_only=True)
    assert "行程校对" in workbook.sheetnames
    assert TASKS[task_id]["state"] == "done"


def test_web_batch_task_snapshot_includes_package_statuses(tmp_path: Path, monkeypatch):
    root = tmp_path / "待整理报销"
    write_file(root / "A" / "invoice.pdf", b"invoice-a")
    write_file(root / "B" / "invoice.pdf", b"invoice-b")
    out_dir = tmp_path / "out"
    task_id = "task-batch-preview"
    TASKS[task_id] = {
        "id": task_id,
        "state": "queued",
        "stage": "排队中",
        "started_at": time.time(),
        "updated_at": time.time(),
        "total": 0,
        "completed": 0,
        "files": [],
        "output_dir": "",
        "excel_path": "",
        "error": "",
    }
    form = {
        "folder": str(root),
        "out_dir": str(out_dir),
        "organize_mode": "batch_subfolders",
        "traveler": "张三",
        "department": "技术部",
        "trip_start_date": "2026-03-01",
        "trip_end_date": "2026-03-02",
        "max_workers": "1",
        "timeout_seconds": "120",
    }
    monkeypatch.setattr(
        "invoice_agent.web.load_agent_config",
        lambda path: type(
            "Config",
            (),
            {
                "ocr_provider": "async_jobs",
                "paddleocr_job_url": "https://example.com/jobs",
                "paddleocr_access_token": "",
                "paddleocr_model": "PaddleOCR-VL-1.6",
                "request_timeout_seconds": 60,
                "retry_max_attempts": 3,
                "retry_base_delay_seconds": 1.0,
                "fallback_api_url": "",
            },
        )(),
    )
    monkeypatch.setattr("invoice_agent.web.SdkOcrProvider", lambda **kwargs: FakeOcrProvider())

    run_organize_from_form(form, task_id=task_id)
    snapshot = get_task_snapshot(task_id)

    assert snapshot["mode"] == "batch_subfolders"
    assert snapshot["state"] == "review"
    assert snapshot["total"] == 2
    assert snapshot["completed"] == 2
    assert [package["name"] for package in snapshot["packages"]] == ["A", "B"]
    assert all(package["state"] == "review" for package in snapshot["packages"])
    assert all(package["can_export"] is True for package in snapshot["packages"])
    assert "_packages" not in snapshot
    assert "_records" not in snapshot
    assert not (out_dir / "A" / "00_报销清单.xlsx").exists()


def test_web_batch_task_publishes_package_preview_before_all_packages_finish(tmp_path: Path, monkeypatch):
    root = tmp_path / "待整理报销"
    write_file(root / "A" / "invoice.pdf", b"invoice-a")
    write_file(root / "B" / "invoice.pdf", b"invoice-b")
    out_dir = tmp_path / "out"
    task_id = "task-batch-streaming-preview"
    TASKS[task_id] = {
        "id": task_id,
        "mode": "batch_subfolders",
        "state": "queued",
        "stage": "排队中",
        "started_at": time.time(),
        "updated_at": time.time(),
        "total": 0,
        "completed": 0,
        "files": [],
        "output_dir": "",
        "excel_path": "",
        "error": "",
    }
    form = {
        "folder": str(root),
        "out_dir": str(out_dir),
        "organize_mode": "batch_subfolders",
        "traveler": "张三",
        "department": "技术部",
        "trip_start_date": "2026-03-01",
        "trip_end_date": "2026-03-02",
        "max_workers": "1",
        "timeout_seconds": "120",
    }
    provider = BlockingSecondPackageProvider()
    monkeypatch.setattr(
        "invoice_agent.web.load_agent_config",
        lambda path: type(
            "Config",
            (),
            {
                "ocr_provider": "async_jobs",
                "paddleocr_job_url": "https://example.com/jobs",
                "paddleocr_access_token": "",
                "paddleocr_model": "PaddleOCR-VL-1.6",
                "request_timeout_seconds": 60,
                "retry_max_attempts": 3,
                "retry_base_delay_seconds": 1.0,
                "fallback_api_url": "",
            },
        )(),
    )
    monkeypatch.setattr("invoice_agent.web.SdkOcrProvider", lambda **kwargs: provider)

    thread = threading.Thread(target=run_organize_from_form, args=(form,), kwargs={"task_id": task_id})
    thread.start()
    assert provider.started_second.wait(timeout=1)
    snapshot = get_task_snapshot(task_id)
    provider.release_second.set()
    thread.join(timeout=2)

    assert snapshot["state"] == "running"
    assert snapshot["mode"] == "batch_subfolders"
    assert snapshot["completed"] == 1
    assert snapshot["packages"][0]["name"] == "A"
    assert snapshot["packages"][0]["state"] == "review"
    assert snapshot["packages"][0]["preview"]["main_rows"][0]["原文件名"] == "invoice.pdf"
    assert snapshot["packages"][0]["can_export"] is True
    assert len(snapshot["packages"]) == 1
    final_snapshot = get_task_snapshot(task_id)
    assert final_snapshot["state"] == "review"
    assert final_snapshot["completed"] == 2


def test_export_single_batch_package_writes_only_that_package(tmp_path: Path):
    record_a = make_record(tmp_path, "a.pdf", "网约车发票", "88.00")
    record_b = make_record(tmp_path, "b.pdf", "网约车发票", "99.00")
    for index, record in enumerate([record_a, record_b], start=1):
        analyze_records([record])
        record.sequence = index
    TASKS["task-batch-export-one"] = {
        "id": "task-batch-export-one",
        "mode": "batch_subfolders",
        "state": "review",
        "stage": "等待确认",
        "started_at": time.time(),
        "updated_at": time.time(),
        "packages": [
            {"id": "pkg-a", "name": "A", "state": "review", "output_dir": str(tmp_path / "out" / "A"), "can_export": True, "preview": build_preview([record_a]), "excel_path": ""},
            {"id": "pkg-b", "name": "B", "state": "review", "output_dir": str(tmp_path / "out" / "B"), "can_export": True, "preview": build_preview([record_b]), "excel_path": ""},
        ],
        "_packages": {
            "pkg-a": {"records": [record_a], "output_dir": tmp_path / "out" / "A", "apply": False, "trip_audit": None},
            "pkg-b": {"records": [record_b], "output_dir": tmp_path / "out" / "B", "apply": False, "trip_audit": None},
        },
    }

    result = export_batch_package("task-batch-export-one", "pkg-a")

    assert result["excel_path"] == str(tmp_path / "out" / "A" / "00_报销清单.xlsx")
    assert result["company_excel_path"] == str(tmp_path / "out" / "A" / "01_公司报销单.xlsx")
    assert (tmp_path / "out" / "A" / "00_报销清单.xlsx").exists()
    assert not (tmp_path / "out" / "B" / "00_报销清单.xlsx").exists()
    packages = {package["id"]: package for package in TASKS["task-batch-export-one"]["packages"]}
    assert packages["pkg-a"]["state"] == "done"
    assert packages["pkg-a"]["can_export"] is False
    assert packages["pkg-b"]["state"] == "review"


def test_export_all_batch_packages_reports_partial_failures(tmp_path: Path):
    record = make_record(tmp_path, "a.pdf", "网约车发票", "88.00")
    TASKS["task-batch-export-all"] = {
        "id": "task-batch-export-all",
        "mode": "batch_subfolders",
        "state": "review",
        "stage": "等待确认",
        "started_at": time.time(),
        "updated_at": time.time(),
        "packages": [
            {"id": "pkg-a", "name": "A", "state": "review", "output_dir": str(tmp_path / "out" / "A"), "can_export": True, "preview": build_preview([record]), "excel_path": ""},
            {"id": "pkg-b", "name": "B", "state": "failed", "output_dir": str(tmp_path / "out" / "B"), "can_export": False, "preview": {}, "error": "Missing trip info", "excel_path": ""},
        ],
        "_packages": {
            "pkg-a": {"records": [record], "output_dir": tmp_path / "out" / "A", "apply": False, "trip_audit": None},
            "pkg-b": {"records": [], "output_dir": tmp_path / "out" / "B", "apply": False, "trip_audit": None},
        },
    }

    result = export_all_batch_packages("task-batch-export-all")

    assert result["exported"] == [{"package_id": "pkg-a", "name": "A", "excel_path": str(tmp_path / "out" / "A" / "00_报销清单.xlsx")}]
    assert result["failed"] == [{"package_id": "pkg-b", "name": "B", "error": "Missing trip info"}]
    assert TASKS["task-batch-export-all"]["state"] == "done"


def test_review_edits_update_preview_raw_results_and_exported_excel(tmp_path: Path):
    out_dir = tmp_path / "out"
    record = make_record(tmp_path, "invoice.pdf", "网约车发票", "88.00", seller_name="滴滴出行科技有限公司")
    analyze_records([record])
    record.sequence = 1
    TASKS["task-edit"] = {
        "id": "task-edit",
        "state": "review",
        "stage": "等待确认",
        "started_at": time.time(),
        "updated_at": time.time(),
        "files": [{"path": str(record.source_path), "name": record.original_name, "status": "已识别"}],
        "preview": build_preview([record]),
        "can_export": True,
        "output_dir": str(out_dir),
        "excel_path": "",
        "_records": [record],
        "_output_dir": out_dir,
        "_apply": False,
    }

    response = update_task_record(
        "task-edit",
        "1",
        {
            "total_with_tax": "99.50",
            "document_date": "2026-03-05",
            "document_type": "住宿发票",
            "reimbursement_category": "住宿费",
            "include_in_amount": "是",
            "invoice_number": "INV-EDIT",
            "seller_name": "人工校对酒店",
            "buyer_name": "杭州勤合能源科技有限公司",
            "origin": "南京",
            "destination": "杭州",
            "description": "住宿校对",
            "risk_note": "",
        },
    )

    edited = TASKS["task-edit"]["_records"][0]
    assert edited.total_with_tax == "99.50"
    assert edited.document_type == "住宿发票"
    assert edited.high_level_category == "住宿费"
    assert "INV-EDIT" in edited.new_name
    assert response["preview"]["review_cards"][0]["价税合计"] == 99.5
    assert response["preview"]["summary_rows"][3] == {"报销大类": "通行费", "计入金额合计": 99.5, "张数": 1}
    assert TASKS["task-edit"]["files"][0]["amount"] == "99.50"
    raw_results = json.loads((out_dir / "raw_results.json").read_text(encoding="utf-8"))
    assert raw_results[0]["total_with_tax"] == "99.50"
    assert raw_results[0]["document_type"] == "住宿发票"

    export_task("task-edit")

    workbook = load_workbook(out_dir / "00_报销清单.xlsx", data_only=True)
    row = next(workbook["报销清单"].iter_rows(min_row=2, values_only=True))
    assert row[6] == "2026-03-05"
    assert row[7] == "住宿发票"
    assert row[8] == "住宿费"
    assert row[10] == "INV-EDIT"
    assert row[11] == "人工校对酒店"
    assert row[15] == 99.5


def test_review_edits_recompute_trip_audit_preview(tmp_path: Path):
    out_dir = tmp_path / "out"
    record = make_record(tmp_path, "didi.pdf", "网约车发票", "40.00", seller_name="滴滴出行科技有限公司")
    record.sequence = 1
    TASKS["task-audit-edit"] = {
        "id": "task-audit-edit",
        "state": "review",
        "stage": "等待确认",
        "started_at": time.time(),
        "updated_at": time.time(),
        "files": [{"path": str(record.source_path), "name": record.original_name, "status": "已识别"}],
        "preview": build_preview([record]),
        "can_export": True,
        "output_dir": str(out_dir),
        "excel_path": "",
        "_records": [record],
        "_output_dir": out_dir,
        "_apply": False,
        "_trip_audit_policy": TripAuditPolicy(city_transport_daily_limit="50"),
    }

    response = update_task_record(
        "task-audit-edit",
        "1",
        {
            "total_with_tax": "88.00",
            "document_date": "2026-03-01",
            "document_type": "网约车发票",
            "include_in_amount": "是",
        },
    )

    assert any(row["校对类别"] == "市内交通" for row in response["preview"]["trip_audit_rows"])


def test_web_review_progress_rows_include_final_risk_messages(tmp_path: Path, monkeypatch):
    src = tmp_path / "上海出差"
    write_file(src / "invoice.pdf", b"invoice")
    write_file(src / "duplicate.pdf", b"invoice")
    out_dir = tmp_path / "out"
    task_id = "task-risk-progress"
    TASKS[task_id] = {
        "id": task_id,
        "state": "queued",
        "stage": "排队中",
        "started_at": time.time(),
        "updated_at": time.time(),
        "total": 0,
        "completed": 0,
        "files": [],
        "output_dir": "",
        "excel_path": "",
        "error": "",
    }
    form = {
        "folder": str(src),
        "out_dir": str(out_dir),
        "traveler": "张三",
        "department": "技术部",
        "trip_start_date": "2026-03-01",
        "trip_end_date": "2026-03-02",
        "max_workers": "1",
        "timeout_seconds": "120",
    }
    monkeypatch.setattr(
        "invoice_agent.web.load_agent_config",
        lambda path: type(
            "Config",
            (),
            {
                "ocr_provider": "async_jobs",
                "paddleocr_job_url": "https://example.com/jobs",
                "paddleocr_access_token": "",
                "paddleocr_model": "PaddleOCR-VL-1.6",
                "request_timeout_seconds": 60,
                "retry_max_attempts": 3,
                "retry_base_delay_seconds": 1.0,
                "fallback_api_url": "",
            },
        )(),
    )
    monkeypatch.setattr("invoice_agent.web.SdkOcrProvider", lambda **kwargs: FakeOcrProvider())

    run_organize_from_form(form, task_id=task_id)

    messages = {row["name"]: row["message"] for row in TASKS[task_id]["files"]}
    assert "Hash重复" in messages["invoice.pdf"]
    assert "Hash重复" in messages["duplicate.pdf"]
    assert "重复发票不计入汇总" in messages["duplicate.pdf"]


def test_choose_path_rejects_unknown_kind():
    result = handle_choose_path("kind=unknown")

    assert "Unsupported chooser kind" in result["error"]
