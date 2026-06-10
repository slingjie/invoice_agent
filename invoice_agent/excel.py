from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .analysis import reimbursement_summary_rows
from .extractor import parse_amount
from .models import ExpenseRecord
from .trip_audit import TripAuditResult


MAIN_HEADERS = [
    "序号",
    "项目名",
    "人员",
    "部门",
    "出差开始日期",
    "出差结束日期",
    "凭证日期",
    "凭证类别",
    "报销大类",
    "是否计入金额",
    "发票号码",
    "销方名称",
    "购方名称",
    "金额不含税",
    "税额",
    "价税合计",
    "起点",
    "终点",
    "行程/住宿说明",
    "原文件名",
    "新文件名",
    "原文件路径",
    "复制后路径",
    "文件Hash",
    "重复标记",
    "识别状态",
    "置信度/风险提示",
    "人工备注",
    "关联组",
]


TRIP_AUDIT_HEADERS = ["校对类别", "风险级别", "结论", "证据", "关联序号", "建议动作"]


def write_workbook(path: Path, records: List[ExpenseRecord], trip_audit: TripAuditResult | None = None) -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "报销清单"
    main.append(MAIN_HEADERS)
    for record in records:
        main.append(_main_row(record))
    main.append(_main_total_row(records))

    summary = workbook.create_sheet("类别汇总")
    summary.append(["报销大类", "计入金额合计", "张数"])
    summary_rows = reimbursement_summary_rows(records)
    for category, total, count in summary_rows:
        summary.append([category, float(total), count])
    total, count = _summary_total(summary_rows)
    summary.append(["合计总金额", float(total), count])

    risks = workbook.create_sheet("重复与风险")
    risks.append(["序号", "原文件名", "凭证类别", "重复标记", "识别状态", "风险提示"])
    for record in records:
        if record.duplicate_mark or record.recognition_status != "已识别" or record.risk_note:
            risks.append(
                [
                    record.sequence,
                    record.original_name,
                    record.document_type,
                    record.duplicate_mark,
                    record.recognition_status,
                    record.risk_note,
                ]
            )
    for item in (trip_audit.items if trip_audit else []):
        if item.severity not in {"warning", "error"}:
            continue
        risks.append(
            [
                ",".join(str(sequence) for sequence in item.related_sequences),
                "",
                "行程校对",
                "",
                item.severity,
                item.conclusion,
            ]
        )

    trip_sheet = workbook.create_sheet("行程校对")
    trip_sheet.append(TRIP_AUDIT_HEADERS)
    if trip_audit:
        for row in _trip_audit_rows(trip_audit):
            trip_sheet.append([row[header] for header in TRIP_AUDIT_HEADERS])
        if trip_audit.llm_review:
            trip_sheet.append(["模型复核", "info", trip_audit.llm_review, "", "", ""])

    rename = workbook.create_sheet("重命名计划")
    rename.append(["序号", "原文件路径", "新文件名", "复制后路径", "是否执行"])
    for record in records:
        rename.append(
            [
                record.sequence,
                str(record.source_path),
                record.new_name,
                record.copied_path,
                "是" if record.copied_path else "否",
            ]
        )

    for sheet in workbook.worksheets:
        _format_sheet(sheet)
    workbook.save(path)


def _main_row(record: ExpenseRecord) -> List[object]:
    return [
        record.sequence,
        record.project_name,
        record.traveler,
        record.department,
        record.trip_start_date,
        record.trip_end_date,
        record.document_date,
        record.document_type,
        record.reimbursement_category if record.include_in_amount else "",
        "是" if record.include_in_amount else "否",
        record.invoice_number,
        record.seller_name,
        record.buyer_name,
        _float_or_blank(record.total_amount),
        _float_or_blank(record.total_tax),
        _float_or_blank(record.total_with_tax),
        record.origin,
        record.destination,
        record.description,
        record.original_name,
        record.new_name,
        str(record.source_path),
        record.copied_path,
        record.file_hash,
        record.duplicate_mark,
        record.recognition_status,
        record.risk_note,
        record.manual_note,
        record.linked_group,
    ]


def build_preview(records: List[ExpenseRecord], trip_audit: TripAuditResult | None = None) -> dict:
    risk_headers = ["序号", "原文件名", "凭证类别", "重复标记", "识别状态", "风险提示"]
    rename_headers = ["序号", "原文件路径", "新文件名", "复制后路径", "是否执行"]
    summary_rows = reimbursement_summary_rows(records)
    summary_total, summary_count = _summary_total(summary_rows)
    return {
        "review_cards": [_review_card(record) for record in records],
        "overview_rows": [_overview_row(record) for record in records],
        "main_rows": [dict(zip(MAIN_HEADERS, _main_row(record))) for record in records]
        + [dict(zip(MAIN_HEADERS, _main_total_row(records)))],
        "summary_rows": [
            {"报销大类": category, "计入金额合计": float(total), "张数": count}
            for category, total, count in summary_rows
        ]
        + [
            {
                "报销大类": "合计总金额",
                "计入金额合计": float(summary_total),
                "张数": summary_count,
            }
        ],
        "risk_rows": [
            dict(
                zip(
                    risk_headers,
                    [
                        record.sequence,
                        record.original_name,
                        record.document_type,
                        record.duplicate_mark,
                        record.recognition_status,
                        record.risk_note,
                    ],
                )
            )
            for record in records
            if record.duplicate_mark or record.recognition_status != "已识别" or record.risk_note
        ]
        + [
            {
                "序号": ",".join(str(sequence) for sequence in item.related_sequences),
                "原文件名": "",
                "凭证类别": "行程校对",
                "重复标记": "",
                "识别状态": item.severity,
                "风险提示": item.conclusion,
            }
            for item in (trip_audit.items if trip_audit else [])
            if item.severity in {"warning", "error"}
        ],
        "trip_audit_rows": _trip_audit_rows(trip_audit) if trip_audit else [],
        "trip_audit_llm_review": trip_audit.llm_review if trip_audit else "",
        "rename_rows": [
            dict(
                zip(
                    rename_headers,
                    [
                        record.sequence,
                        str(record.source_path),
                        record.new_name,
                        record.copied_path,
                        "是" if record.copied_path else "否",
                    ],
                )
            )
            for record in records
        ],
    }


def _trip_audit_rows(trip_audit: TripAuditResult) -> List[dict]:
    return [
        {
            "校对类别": item.category,
            "风险级别": item.severity,
            "结论": item.conclusion,
            "证据": "；".join(item.evidence),
            "关联序号": ",".join(str(sequence) for sequence in item.related_sequences),
            "建议动作": item.suggested_action,
        }
        for item in trip_audit.items
    ]


def _review_card(record: ExpenseRecord) -> dict:
    return {
        "序号": record.sequence,
        "原文件名": record.original_name,
        "凭证日期": record.document_date,
        "凭证类别": record.document_type,
        "报销大类": record.reimbursement_category if record.include_in_amount else "",
        "价税合计": _float_or_blank(record.total_with_tax),
        "是否计入金额": "是" if record.include_in_amount else "否",
        "发票号码": record.invoice_number,
        "销方名称": record.seller_name,
        "购方名称": record.buyer_name,
        "起点": record.origin,
        "终点": record.destination,
        "行程/住宿说明": record.description,
        "风险提示": record.risk_note,
    }


def _overview_row(record: ExpenseRecord) -> dict:
    return {
        "序号": record.sequence,
        "日期": record.document_date,
        "凭证类别": record.document_type,
        "报销大类": record.reimbursement_category if record.include_in_amount else "",
        "金额": _float_or_blank(record.total_with_tax),
        "是否计入": "是" if record.include_in_amount else "否",
        "风险": record.risk_note,
        "原文件名": record.original_name,
    }


def _main_total_row(records: List[ExpenseRecord]) -> List[object]:
    total, _ = _included_total(records)
    row: List[object] = [""] * len(MAIN_HEADERS)
    row[0] = "合计总金额"
    row[15] = float(total)
    return row


def _included_total(records: List[ExpenseRecord]):
    total = Decimal("0.00")
    count = 0
    for record in records:
        if not record.include_in_amount:
            continue
        total += parse_amount(record.total_with_tax)
        count += 1
    return total, count


def _summary_total(summary_rows):
    total = Decimal("0.00")
    count = 0
    for _, row_total, row_count in summary_rows:
        total += row_total
        count += row_count
    return total, count


def _float_or_blank(value: str):
    amount = parse_amount(value)
    return "" if amount == Decimal("0") and not value else float(amount)


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        joined = " ".join(str(value or "") for value in values)
        fill = None
        if "无法识别" in joined or "失败" in joined:
            fill = error_fill
        elif "重复" in joined or "待人工确认" in joined or "金额不一致" in joined:
            fill = warning_fill
        if values and values[0] == "合计总金额":
            fill = header_fill
            for cell in row:
                cell.font = Font(bold=True)
        if fill:
            for cell in row:
                cell.fill = fill
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:50])
        sheet.column_dimensions[letter].width = max(10, min(max_len + 2, 40))
