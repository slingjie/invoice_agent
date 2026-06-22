from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from .analysis import meal_allowance_total
from .extractor import parse_amount
from .models import ExpenseRecord


@dataclass(frozen=True)
class CompanyTravelRow:
    kind: str
    date_text: str
    origin: str
    destination: str
    transport: str
    amount: Decimal
    count: int = 1


@dataclass(frozen=True)
class CompanyDailyRow:
    date_text: str
    content: str
    amount: Decimal


@dataclass(frozen=True)
class CompanyDetailLine:
    date_text: str
    person: str
    location: str
    purpose: str
    amount: Decimal
    count: int


@dataclass(frozen=True)
class CompanyReimbursementData:
    project_name: str
    traveler: str
    department: str
    trip_start_date: str
    trip_end_date: str
    trip_days: int
    export_date: date
    travel_rows: list[CompanyTravelRow] = field(default_factory=list)
    daily_rows: list[CompanyDailyRow] = field(default_factory=list)
    detail_sections: dict[str, list[CompanyDetailLine]] = field(default_factory=dict)
    meal_allowance: CompanyDetailLine = field(
        default_factory=lambda: CompanyDetailLine("", "", "", "餐补", Decimal("0.00"), 0)
    )
    lodging_total: Decimal = Decimal("0.00")
    toll_total: Decimal = Decimal("0.00")
    fuel_total: Decimal = Decimal("0.00")
    refund_total: Decimal = Decimal("0.00")
    total: Decimal = Decimal("0.00")
    warnings: list[str] = field(default_factory=list)


def amount_to_chinese_upper(amount: Decimal) -> str:
    value = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if value < 0:
        raise ValueError("金额不能为负数")
    fen_total = int(value * 100)
    yuan, fraction = divmod(fen_total, 100)
    jiao, fen = divmod(fraction, 10)
    parts = ["人民币"]
    if yuan:
        parts.append(_integer_to_chinese(yuan))
        parts.append("元")
    if jiao:
        parts.append("零" if yuan and yuan % 10 == 0 else "")
        parts.append("壹贰叁肆伍陆柒捌玖"[jiao - 1])
        parts.append("角")
    if fen:
        if yuan and not jiao:
            parts.append("零")
        parts.append("壹贰叁肆伍陆柒捌玖"[fen - 1])
        parts.append("分")
    if not jiao and not fen:
        if not yuan:
            parts.append("零元")
        parts.append("整")
    return "".join(parts)


def _integer_to_chinese(value: int) -> str:
    digits = "零壹贰叁肆伍陆柒捌玖"
    small_units = ["", "拾", "佰", "仟"]
    large_units = ["", "万", "亿", "兆"]
    groups: list[int] = []
    while value:
        groups.append(value % 10000)
        value //= 10000
    result = ""
    pending_zero = False
    for index in range(len(groups) - 1, -1, -1):
        group = groups[index]
        if group == 0:
            pending_zero = bool(result)
            continue
        if result and (pending_zero or group < 1000):
            result += "零"
        group_text = ""
        zero_inside = False
        for position in range(3, -1, -1):
            divisor = 10**position
            digit = group // divisor % 10
            if digit:
                if zero_inside and group_text:
                    group_text += "零"
                group_text += digits[digit] + small_units[position]
                zero_inside = False
            elif group_text:
                zero_inside = True
        result += group_text + large_units[index]
        pending_zero = False
    return result or "零"


def build_company_reimbursement_data(
    records: Iterable[ExpenseRecord],
    export_date: date | None = None,
) -> CompanyReimbursementData:
    records = list(records)
    first = records[0] if records else None
    export_date = export_date or date.today()
    project_name = first.project_name if first else ""
    traveler = first.traveler if first else ""
    department = first.department if first else ""
    trip_start = first.trip_start_date if first else ""
    trip_end = first.trip_end_date if first else ""
    trip_days = _inclusive_days(trip_start, trip_end)
    warnings = _base_warnings(project_name, traveler, department)

    travel_rows: list[CompanyTravelRow] = []
    daily_rows: list[CompanyDailyRow] = []
    detail_sections: dict[str, list[CompanyDetailLine]] = {}
    lodging_total = Decimal("0.00")
    toll_total = Decimal("0.00")
    fuel_total = Decimal("0.00")
    refund_total = Decimal("0.00")
    invoice_total = Decimal("0.00")

    for record in records:
        if not record.include_in_amount or record.document_type == "行程单":
            continue
        amount = parse_amount(record.total_with_tax)
        invoice_total += amount
        fine = record.reimbursement_category
        detail_name = _detail_section(record)
        detail_sections.setdefault(detail_name, []).append(_detail_line(record, amount))
        if fine == "行程交通费":
            travel_rows.append(_travel_row(record, amount, "intercity"))
        elif fine in {"通行费", "市区交通费"}:
            travel_rows.append(_travel_row(record, amount, "city"))
        elif fine == "住宿费":
            lodging_total += amount
        elif fine == "过路费":
            toll_total += amount
        elif fine == "油费":
            fuel_total += amount
        elif fine == "退改费":
            refund_total += amount
        else:
            daily_rows.append(
                CompanyDailyRow(
                    date_text=record.document_date,
                    content=_clean_text(record.manual_note or record.description or record.seller_name or record.document_type),
                    amount=amount,
                )
            )

    meal_total, meal_days = meal_allowance_total(records)
    meal_line = CompanyDetailLine(
        date_text=_date_range_text(trip_start, trip_end),
        person=traveler,
        location=_trip_destination(records),
        purpose="餐补",
        amount=meal_total,
        count=0,
    )
    if meal_total:
        detail_sections.setdefault("差旅费", []).append(meal_line)
    total = invoice_total + meal_total
    return CompanyReimbursementData(
        project_name=project_name,
        traveler=traveler,
        department=department,
        trip_start_date=trip_start,
        trip_end_date=trip_end,
        trip_days=meal_days or trip_days,
        export_date=export_date,
        travel_rows=travel_rows,
        daily_rows=daily_rows,
        detail_sections=detail_sections,
        meal_allowance=meal_line,
        lodging_total=lodging_total,
        toll_total=toll_total,
        fuel_total=fuel_total,
        refund_total=refund_total,
        total=total,
        warnings=warnings,
    )


def _travel_row(record: ExpenseRecord, amount: Decimal, kind: str) -> CompanyTravelRow:
    transport = "火车" if record.document_type == "高铁发票" else ("公共交通" if kind == "city" else record.document_type)
    return CompanyTravelRow(
        kind=kind,
        date_text=record.document_date,
        origin=_clean_text(record.origin),
        destination=_clean_text(record.destination),
        transport=transport,
        amount=amount,
    )


def _detail_line(record: ExpenseRecord, amount: Decimal) -> CompanyDetailLine:
    if record.reimbursement_category == "住宿费":
        date_text = _date_range_text(record.trip_start_date, record.trip_end_date)
        purpose = "住宿费"
    else:
        date_text = record.document_date
        purpose = _clean_text(record.manual_note or record.description or record.document_type)
    return CompanyDetailLine(
        date_text=date_text,
        person=record.traveler,
        location=_clean_text(record.seller_name),
        purpose=purpose,
        amount=amount,
        count=1,
    )


def _detail_section(record: ExpenseRecord) -> str:
    if record.reimbursement_category == "住宿费":
        return "差旅费"
    if record.high_level_category == "招待费":
        return "招待费（餐饮、娱乐）"
    if record.high_level_category == "材料费":
        return "材料"
    if record.high_level_category == "办公费":
        return "办公费"
    if _is_gift(record):
        return "礼品、礼卡"
    if record.reimbursement_category in {"行程交通费", "通行费", "市区交通费", "过路费", "油费", "退改费"}:
        return "交通费"
    return "其他费用"


def _is_gift(record: ExpenseRecord) -> bool:
    text = " ".join([record.manual_note, record.description, record.seller_name, record.raw_text])
    return any(keyword in text for keyword in ["赠送", "礼品", "礼卡", "客户维护"])


def _base_warnings(project_name: str, traveler: str, department: str) -> list[str]:
    warnings = []
    for label, value in [("项目名", project_name), ("报销人", traveler), ("部门", department)]:
        if not value or value.strip() in {"1", "2", "3"}:
            warnings.append(f"{label}缺失或疑似占位值，请在公司报销单中人工确认")
    return warnings


def _inclusive_days(start_text: str, end_text: str) -> int:
    try:
        start = datetime.strptime(start_text, "%Y-%m-%d").date()
        end = datetime.strptime(end_text, "%Y-%m-%d").date()
    except ValueError:
        return 0
    return max((end - start).days + 1, 0)


def _date_range_text(start_text: str, end_text: str) -> str:
    if start_text and end_text:
        return f"{start_text}至{end_text}"
    return start_text or end_text


def _trip_destination(records: Iterable[ExpenseRecord]) -> str:
    for record in records:
        if record.reimbursement_category == "行程交通费" and record.destination:
            return _clean_text(record.destination)
    return ""


def _clean_text(value: str) -> str:
    return " ".join((value or "").replace("\\n", " ").replace("\r", " ").replace("\n", " ").split())


def company_template_path() -> Path:
    return Path(__file__).with_name("templates") / "company_reimbursement_template.xlsx"


def write_company_workbook(
    path: Path,
    data: CompanyReimbursementData,
    template_path: Path | None = None,
) -> None:
    template = template_path or company_template_path()
    workbook = load_workbook(template, data_only=False)
    _fill_detail_sheet(workbook["报销明细表"], data)
    _fill_travel_sheets(workbook, data)
    _fill_daily_sheets(workbook, data)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


def _fill_detail_sheet(sheet, data: CompanyReimbursementData) -> None:
    sheet["B2"] = data.project_name
    sheet["E2"] = data.export_date
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= 4:
            sheet.unmerge_cells(str(merged_range))
    for row in range(4, sheet.max_row + 1):
        for column in range(1, 7):
            sheet.cell(row, column).value = None
    sections = [
        "招待费（餐饮、娱乐）",
        "差旅费",
        "交通费",
        "办公费",
        "礼品、礼卡",
        "材料",
        "其他费用",
    ]
    rows: list[tuple[str, CompanyDetailLine | None]] = []
    for section in sections:
        values = data.detail_sections.get(section, [])
        if not values and section == "其他费用":
            continue
        rows.append((section, None))
        rows.extend((section, value) for value in values)
        rows.append(("小计", None))
    start_row = 4
    current = start_row
    section_start = 0
    total_cells: list[str] = []
    for label, line in rows:
        if line is None and label != "小计":
            _copy_row_style(sheet, 4, current, 6)
            sheet.merge_cells(start_row=current, start_column=1, end_row=current, end_column=6)
            sheet.cell(current, 1).value = label
            section_start = current + 1
        elif label == "小计":
            _copy_row_style(sheet, 7, current, 6)
            sheet.cell(current, 1).value = "小计"
            if current > section_start:
                sheet.cell(current, 5).value = f"=SUM(E{section_start}:E{current - 1})"
                sheet.cell(current, 6).value = f"=SUM(F{section_start}:F{current - 1})"
            else:
                sheet.cell(current, 5).value = 0
                sheet.cell(current, 6).value = 0
            total_cells.append(f"E{current}")
        else:
            _copy_row_style(sheet, 5, current, 6)
            values = [line.date_text, line.person, line.location, line.purpose, float(line.amount), line.count]
            for column, value in enumerate(values, start=1):
                sheet.cell(current, column).value = value
        current += 1
    _copy_row_style(sheet, 44, current, 6)
    sheet.cell(current, 1).value = "总计"
    sheet.cell(current, 5).value = "=" + "+".join(total_cells) if total_cells else 0
    sheet.cell(current, 6).value = f"=SUM(F{start_row}:F{current - 1})"
    footer_row = current + 2
    _copy_row_style(sheet, 45, footer_row, 6)
    sheet.cell(footer_row, 1).value = "费用超支情况说明"
    signature_row = current + 5
    _copy_row_style(sheet, 48, signature_row, 6)
    sheet.cell(signature_row, 1).value = "报销人："
    sheet.cell(signature_row, 2).value = data.traveler
    sheet.cell(signature_row, 3).value = "项目经理："
    sheet.cell(signature_row, 4).value = "财务审核："
    sheet.cell(signature_row, 5).value = "总经理："
    sheet.print_area = f"A1:F{signature_row + 2}"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _fill_travel_sheets(workbook, data: CompanyReimbursementData) -> None:
    base = workbook["差旅费报销单"]
    chunks = _chunks(data.travel_rows, 12) or [[]]
    sheets = [base]
    for index in range(1, len(chunks)):
        copied = workbook.copy_worksheet(base)
        copied.title = f"差旅费报销单 ({index + 1})"
        sheets.append(copied)
    for index, (sheet, rows) in enumerate(zip(sheets, chunks)):
        _clear_travel_inputs(sheet)
        sheet["J2"] = data.export_date
        sheet["C3"] = data.department
        sheet["E3"] = data.traveler
        sheet["H3"] = data.trip_days
        sheet["J3"] = data.project_name
        sheet["C21"] = data.traveler
        for offset, row in enumerate(rows, start=6):
            sheet.cell(offset, 2).value = row.date_text
            sheet.cell(offset, 3).value = row.origin
            sheet.cell(offset, 4).value = row.destination
            sheet.cell(offset, 5).value = row.transport
            sheet.cell(offset, 6).value = row.count
            sheet.cell(offset, 7).value = float(row.amount)
        page_total = sum((row.amount for row in rows), Decimal("0.00"))
        sheet["G18"] = float(page_total)
        if index == 0:
            sheet["H6"] = "餐补"
            sheet["I6"] = data.trip_days
            sheet["J6"] = float(data.meal_allowance.amount)
            sheet["K6"] = "住宿费"
            sheet["L6"] = float(data.lodging_total)
            sheet["K7"] = "过路费"
            sheet["L7"] = float(data.toll_total)
            sheet["K8"] = "油  费"
            sheet["L8"] = float(data.fuel_total)
            sheet["K9"] = "退改费"
            sheet["L9"] = float(data.refund_total)
            sheet["J18"] = float(data.meal_allowance.amount)
            sheet["L18"] = float(data.lodging_total + data.toll_total + data.fuel_total + data.refund_total)
            sheet_total = (
                page_total
                + data.meal_allowance.amount
                + data.lodging_total
                + data.toll_total
                + data.fuel_total
                + data.refund_total
            )
            sheet["K19"] = float(sheet_total)
            sheet["C19"] = amount_to_chinese_upper(sheet_total)
        else:
            sheet["J18"] = 0
            sheet["L18"] = 0
            sheet["K19"] = float(page_total)
            sheet["C19"] = amount_to_chinese_upper(page_total)
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A5
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _fill_daily_sheets(workbook, data: CompanyReimbursementData) -> None:
    base = workbook["日常费用报销单"]
    if not data.daily_rows:
        workbook.remove(base)
        return
    chunks = _chunks(data.daily_rows, 8)
    sheets = [base]
    for index in range(1, len(chunks)):
        copied = workbook.copy_worksheet(base)
        copied.title = f"日常费用报销单 ({index + 1})"
        sheets.append(copied)
    for sheet, rows in zip(sheets, chunks):
        for row_number in range(5, 13):
            sheet.cell(row_number, 1).value = None
            sheet.cell(row_number, 2).value = None
            sheet.cell(row_number, 10).value = None
        sheet["B3"] = data.department
        sheet["F3"] = data.traveler
        sheet["I3"] = data.export_date
        sheet["A15"] = f"报销人：{data.traveler}"
        for offset, row in enumerate(rows, start=5):
            sheet.cell(offset, 1).value = row.date_text
            sheet.cell(offset, 2).value = row.content
            sheet.cell(offset, 10).value = float(row.amount)
        page_total = sum((row.amount for row in rows), Decimal("0.00"))
        sheet["J13"] = float(page_total)
        sheet["B13"] = amount_to_chinese_upper(page_total)
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A5
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _clear_travel_inputs(sheet) -> None:
    for row in range(6, 18):
        for column in range(2, 8):
            sheet.cell(row, column).value = None
    for cell in ["H6", "I6", "J6", "K6", "L6", "H7", "I7", "J7", "K7", "L7", "K8", "L8", "K9", "L9"]:
        sheet[cell] = None


def _copy_row_style(sheet, source_row: int, target_row: int, max_column: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)


def _chunks(values: list, size: int) -> list[list]:
    return [values[index : index + size] for index in range(0, len(values), size)]
